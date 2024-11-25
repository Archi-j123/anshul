import os
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Define the directory where the files are located
download_dir = "C:\\Users\\krishankanty\\Downloads"

# Get the current date and time
current_time = datetime.now().strftime("%d-%m-%Y_%H-%M-%S")

# Create the new Excel file name
output_file = f"TAP_command_{current_time}.xlsx"

# List of Excel files to include as sheets
file_names = [
    "block_load_profile_interval",
    "block_load",
    "last_token_recharge_amt",
    "total_amt_last_recharge",
    "cur_balance_amount",
    "PCP_value",
    "daily_load",
    "cumm_tamper_count",
    "cur_balance_time",
    "firmware_version",
    "profile_instant",
    "MD_kW",
    "last_token_recharge_time",
    "voltage",
    "metering_mode",
    "payment_mode",
    "nameplate",
    "power_event",
    
    
]

# Initialize a Pandas ExcelWriter
output_path = os.path.join(download_dir, output_file)
valid_files = []  # To track added sheets

with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
    for file_name in file_names:
        file_path = os.path.join(download_dir, f"{file_name}.xlsx")  # Assuming files have `.xlsx` extension
        if os.path.exists(file_path):  # Check if the file exists
            try:
                # Read the file into a DataFrame
                df = pd.read_excel(file_path)
                # Add the DataFrame as a sheet in the new Excel file
                sheet_name = os.path.splitext(file_name)[0]  # Remove the .xlsx extension for the sheet name
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                valid_files.append(sheet_name)
            except Exception as e:
                print(f"Error processing file {file_path}: {e}")
        # else:
        #     print(f"File not found: {file_name}")

# Check if any valid files were added
if not valid_files:
    print("No valid files were found. Exiting.")
    exit(1)

# Apply conditional formatting
workbook = load_workbook(output_path)
green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green for PASS
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red for FAIL

for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            if cell.value == "PASS":
                cell.fill = green_fill
            elif cell.value == "FAIL":
                cell.fill = red_fill

# Save the workbook
workbook.save(output_path)
print(f"Excel file created with conditional formatting: {output_path}")


