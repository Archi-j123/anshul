import pyautogui
import time
import os
import pygetwindow as gw
import pandas as pd
import pyperclip

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from guru.final_result_calculator import add_final_result_column

# # ***************************************************** Clean The txt_clear_previous_data *************************************************

exec(open('guru\\txt_clear_previous_data.py').read()) 

# # ************************************************************* End ... *************************************************************

# # ***************************************************** data prepration  *************************************************

exec(open('C:\\Users\\krishankanty\\Desktop\\anshul\\data_preparation\\block_load_profile_interval.py').read()) 

# # ************************************************************* End ... *************************************************************

# # ************************************************************* guruX ... *************************************************************

# Path to the software executable
software_path = r"C:\\Users\\krishankanty\\Desktop\\GXDLMSDirector.appref-ms"  # Replace with actual executable

# Full path to the file you want to open
file_to_open = r"usin.gxc"

# Open the software
os.startfile(software_path)

# Wait for the software to launch
time.sleep(2)

# Ensure the window is active and in focus
windows = gw.getWindowsWithTitle('GXDLMSDirector')  # Window title should match the actual software title
if windows:
    window = windows[0]
    window.activate()  # Bring the window to the front
    window.maximize()  # Maximize the window
else:
    print("GXDLMSDirector window not found.")
    exit()

# Simulate 'Alt' + 'F' to open the 'File' menu
pyautogui.hotkey('alt', 'f')
time.sleep(1)

# Simulate pressing 'O' to select the "Open" option
pyautogui.press('o')
time.sleep(2)

# Write the full path to the file
pyautogui.write(file_to_open)
time.sleep(1)

# Press 'Enter' to open the file
pyautogui.press('enter')

time.sleep(1)

# Click main expand
pyautogui.click(x=17, y=85)
time.sleep(1)

# Click sub expand 
pyautogui.click(x=36, y=101)
time.sleep(1)

# Click data  
pyautogui.click(x=56, y=117)
time.sleep(1)

# Click profile capture Period
pyautogui.click(x=167, y=691)
time.sleep(1)


# Click connect
pyautogui.click(x=93, y=57)
time.sleep(10)

# Click for read
pyautogui.hotkey('ctrl', 'r')
time.sleep(3)

# Click value
pyautogui.click(x=516, y=189)
time.sleep(1)

# Click for copy row
pyautogui.hotkey('ctrl', 'a')
time.sleep(2)

# Click for copy row
pyautogui.hotkey('ctrl', 'c')
time.sleep(2)

# Get the copied data from clipboard
copied_data = pyperclip.paste()

# Path for the text file in the Downloads folder
text_file_path = os.path.join("data_guruX.txt")

# Append data to the text file
with open(text_file_path, "a") as file:
    # Append CSV data
    # file.write("time_set.csv\n")
    # Append clipboard data
    file.write(copied_data + "\n")

print("CSV and clipboard data appended and saved to data.txt in folder.")

# Close the application
pyautogui.click(x=1574, y=4)
time.sleep(2)

# ****************************************************************** End ***********************************************

# ************************************************************* Creating the DATA Format ... *****************************************

# File paths
input_file_path = 'C:\\Users\\krishankanty\\Desktop\\anshul\\data_guruX.txt'
data_file_path = 'C:\\Users\\krishankanty\\Desktop\\anshul\\data.xlsx'
output_file_path = os.path.join(os.getenv('USERPROFILE'), 'Downloads', 'PCP_value.xlsx')

# Step 1: Read and parse data from data_guruX.txt
with open(input_file_path, 'r') as file:
    data = file.read()

# Split into rows and columns
rows = data.strip().split("\n")
parsed_data = [row.split() for row in rows]

# Define headers for data_guruX.txt
headers_file1 = ['PCP_value']

# Create DataFrame for data_guruX.txt
df_file1 = pd.DataFrame(parsed_data, columns=headers_file1)

# Convert relevant columns to numeric
for col in ['PCP_value']:
    df_file1[col] = pd.to_numeric(df_file1[col], errors='coerce').fillna(0).astype(float)

# Add prefix 'file1_' to distinguish from other data sources
df_file1 = df_file1.add_prefix('file1_')

# Step 2: Load data from data.xlsx and select specific columns
columns_file2 = ['date_time', 'meter_ip_address', 'command_name', 'PCP_value', 'data_type']
df_file2 = pd.read_excel(data_file_path, usecols=columns_file2).add_prefix('file2_')

# Step 3: Merge DataFrames, keeping all columns from both DataFrames
df_merged = pd.concat([df_file1, df_file2], axis=1)

# Step 4: Compare columns with a 1% tolerance and create Result columns
comparison_columns = ['PCP_value']
for col in comparison_columns:
    result_column = f"{col}_result"
    df_merged[result_column] = df_merged.apply(
        lambda row, col=col: 'PASS' if abs(row[f'file1_{col}'] - row[f'file2_{col}']) <= 0.01 * row[f'file1_{col}'] else 'FAIL',
        axis=1
    )

df_merged = add_final_result_column(df_merged)

# Step 5: Save the merged DataFrame to profile_instant.xlsx
df_merged.to_excel(output_file_path, index=False)

# Step 6: Apply conditional formatting to highlight 'PASS' in green and 'FAIL' in red
# Load workbook and select the active sheet
wb = load_workbook(output_file_path)
ws = wb.active

# Define fill colors
green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

# Apply formatting based on value in each result column
for col in range(len(df_file1.columns) + len(df_file2.columns) + 1, ws.max_column + 1):  # Result columns start after original columns
    for row in range(2, ws.max_row + 1):  # Skip header row
        cell = ws.cell(row=row, column=col)
        if cell.value == "PASS":
            cell.fill = green_fill
        elif cell.value == "FAIL":
            cell.fill = red_fill

# Save the workbook with formatting
wb.save(output_file_path)

# Print the path where the file is saved
print(f"File saved to: {output_file_path}")


# ************************************************************* End ... *********************************************************************




