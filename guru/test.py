import os
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
from datetime import datetime

# ************************************************************* Creating the DATA Format ... *****************************************
# Paths
input_file_path = 'C:\\Users\\krishankanty\\Desktop\\anshul\\data_guruX.txt'
data_file_path = 'C:\\Users\\krishankanty\\Desktop\\anshul\\data.xlsx'
output_file_path = os.path.join(os.getenv('USERPROFILE'), 'Downloads', 'cur_balance_time.xlsx')

# Step 1: Read and parse data from data_guruX.txt
with open(input_file_path, 'r') as file:
    data = file.read()

# Split into rows and columns
rows = data.strip().split("\n")
parsed_data = [row.split() for row in rows]

# Define headers for data_guruX.txt
headers_file1 = ['cur_balance_date', 'cur_balance_time', 'AM/PM']

# Create DataFrame for data_guruX.txt
df_file1 = pd.DataFrame(parsed_data, columns=headers_file1)

# Convert date and time fields in df_file1 to datetime format
df_file1['cur_balance_date'] = pd.to_datetime(df_file1['cur_balance_date'], format='%d-%m-%Y', errors='coerce')
df_file1['cur_balance_time'] = pd.to_datetime(df_file1['cur_balance_time'], format='%H:%M:%S', errors='coerce').dt.time

# Add prefix to distinguish columns
df_file1 = df_file1.add_prefix('file1_')

# Save df_file1 to Excel
df_file1.to_excel(output_file_path, index=False)
print(f"Data successfully saved to {output_file_path}")

# Step 2: Load data from data.xlsx and select specific columns
columns_file2 = ['date_time', 'meter_ip_address', 'command_name', 'cur_balance_date', 'cur_balance_time', 'data_type']
df_file2 = pd.read_excel(data_file_path, usecols=columns_file2).add_prefix('file2_')

# Convert date and time fields in df_file2 to datetime format
df_file2['file2_cur_balance_date'] = pd.to_datetime(df_file2['file2_cur_balance_date'], format='%d-%m-%Y', errors='coerce')
df_file2['file2_cur_balance_time'] = pd.to_datetime(df_file2['file2_cur_balance_time'], format='%H:%M:%S', errors='coerce').dt.time

# Step 3: Merge DataFrames
df_merged = pd.concat([df_file1, df_file2], axis=1)

# Step 4: Compare columns and create Result columns
comparison_columns = ['cur_balance_date', 'cur_balance_time']
for col in comparison_columns:
    result_column = f"{col}_result"
    df_merged[result_column] = df_merged.apply(
        lambda row, col=col: 'PASS' if (
            pd.notna(row[f'file1_{col}']) and 
            pd.notna(row[f'file2_{col}']) and 
            row[f'file1_{col}'] == row[f'file2_{col}']
        ) else 'FAIL',
        axis=1
    )

# df_merged = add_final_result_column(df_merged)

# Step 5: Save merged DataFrame to output file
df_merged.to_excel(output_file_path, index=False)

# Step 6: Apply conditional formatting
wb = load_workbook(output_file_path)
ws = wb.active

green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

for col in range(len(df_file1.columns) + len(df_file2.columns) + 1, ws.max_column + 1):  
    for row in range(2, ws.max_row + 1):  
        cell = ws.cell(row=row, column=col)
        if cell.value == "PASS":
            cell.fill = green_fill
        elif cell.value == "FAIL":
            cell.fill = red_fill

# Save the workbook
wb.save(output_file_path)
print(f"File saved to: {output_file_path}")


# ************************************************************* End ... *********************************************************************

